"""Import de las 37 ordenes historicas de "Ordenes de Produccion.xlsx" como
production_runs reales (folio OP-2026-0001..0037). Ver
docs/superpowers/specs/2026-08-04-certificados-historicos-design.md.

Uso:
    python -m backend.scripts.import_historical_orders \
        --xlsx "C:/Users/MSI I7/Desktop/Trabajo/Joyeria/Ordenes de Producción.xlsx" \
        --created-by-username <username> \
        [--commit]

Sin --commit corre en modo dry-run: parsea, resuelve usuario, imprime el
resumen, no escribe nada en la base.

Estas ordenes son solo el registro de las actas de papel: no referencian
ninguna materia prima real del inventario (raw_material_item_id queda NULL)
ni generan movimiento de inventario alguno — ver decision #6 del spec.
"""
from __future__ import annotations

import argparse
import datetime as dt
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path
from uuid import UUID, uuid4

import openpyxl

from backend.modules.auth.models import AuthUser
from backend.modules.database.session import SessionLocal
from backend.modules.production.models import (
    ProductionProcess,
    ProductionProcessStage,
    ProductionRun,
    ProductionRunEventLine,
    ProductionRunStage,
    ProductionRunStageStatus,
    ProductionRunStatus,
)

# ProductionRun.target_product_type_id tiene un ForeignKey a product_types.id.
# Ese modelo vive en otro modulo y nadie mas lo importa en este script (a
# diferencia de backend/app/main.py, que si lo hace). Sin este import,
# SQLAlchemy no puede resolver esa FK al armar el orden de insercion y
# session.flush()/commit() de una ProductionRun revienta con
# NoReferencedTableError apenas se corre con --commit. Import solo por su
# efecto de registrar la tabla en Base.metadata.
from backend.modules.product_types import models as _product_types_models  # noqa: F401

SHEETS = ("1-18", "19-37")
PROCESS_NAME = "Producción histórica migrada"


@dataclass
class EventLine:
    gramos: Decimal
    detalle: str | None


@dataclass
class Event:
    fecha: dt.date | None
    lines: list[EventLine] = field(default_factory=list)


@dataclass
class HistoricalOrder:
    order_id: int
    order_name: str | None
    responsable: str | None
    material: str | None
    entrega_events: list[Event] = field(default_factory=list)
    recibido_events: list[Event] = field(default_factory=list)


def _parse_side_events(rows: list[tuple], start_index: int) -> tuple[list[Event], int]:
    """`start_index` apunta a la fila 'Entregado'/'Recibido'; la fila
    start_index+1 es el encabezado 'Fecha'. Devuelve los eventos (un grupo
    por cada fecha real, blancos se pegan al grupo anterior) y el indice de
    la siguiente fila sin consumir."""
    events: list[Event] = []
    j = start_index + 2
    while j < len(rows):
        row = rows[j]
        if row[0] in ("Tipo", "ID"):
            break
        cell = row[0]
        gramos_raw = row[1]
        detalle = row[2] if len(row) > 2 else None
        if isinstance(cell, dt.datetime):
            events.append(Event(fecha=cell.date()))
        if gramos_raw is not None and events:
            try:
                gramos = Decimal(str(gramos_raw))
            except Exception:
                gramos = None
            if gramos is not None:
                events[-1].lines.append(EventLine(gramos=gramos, detalle=str(detalle).strip() if detalle else None))
        elif gramos_raw is not None and not events:
            # Fila con gramos antes de cualquier fecha (no deberia pasar en
            # el archivo real, pero por si acaso arranca un grupo sin fecha).
            events.append(Event(fecha=None))
            try:
                events[-1].lines.append(EventLine(gramos=Decimal(str(gramos_raw)), detalle=str(detalle).strip() if detalle else None))
            except Exception:
                pass
        j += 1
    return events, j


def parse_orders(xlsx_path: Path) -> list[HistoricalOrder]:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    orders: list[HistoricalOrder] = []
    for sheet_name in SHEETS:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
        i = 0
        while i < len(rows):
            row = rows[i]
            if row[0] == "ID":
                header = rows[i + 1]
                order = HistoricalOrder(
                    order_id=int(header[0]),
                    order_name=(str(header[1]).strip() if header[1] else None),
                    responsable=(str(header[2]).strip() if header[2] else None),
                    material=(str(header[3]).strip() if header[3] else None),
                )
                j = i + 2
                while j < len(rows) and rows[j][0] != "ID":
                    if rows[j][0] == "Entregado":
                        order.entrega_events, j = _parse_side_events(rows, j)
                        continue
                    if rows[j][0] == "Recibido":
                        order.recibido_events, j = _parse_side_events(rows, j)
                        continue
                    j += 1
                orders.append(order)
                i = j
            else:
                i += 1
    orders.sort(key=lambda o: o.order_id)
    return orders


def _resolve_user(session, username: str) -> AuthUser:
    from sqlalchemy import select

    user = session.execute(select(AuthUser).where(AuthUser.username == username)).scalar_one_or_none()
    if user is None:
        raise SystemExit(f"No existe ningun usuario con username '{username}'.")
    return user


def _get_or_create_process(session) -> ProductionProcess:
    from sqlalchemy import select

    existing = session.execute(
        select(ProductionProcess).where(ProductionProcess.name == PROCESS_NAME)
    ).scalar_one_or_none()
    if existing is not None:
        return existing
    process = ProductionProcess(
        name=PROCESS_NAME,
        description="Proceso generico para las ordenes migradas del Excel historico de papel.",
        waste_limit_percent=Decimal("100"),
        is_active=False,
        stages=[
            ProductionProcessStage(name="Entregado", stage_type="PROCESS", stage_order=1, is_active=True),
            ProductionProcessStage(name="Recibido", stage_type="PROCESS", stage_order=2, is_active=True),
        ],
    )
    session.add(process)
    session.flush()
    return process


def _next_folio_numbers(count: int) -> list[str]:
    return [f"OP-2026-{n:04d}" for n in range(1, count + 1)]


def _abort_if_already_imported(session) -> None:
    """production_code/root_production_code no tienen constraint UNIQUE en
    la base (solo index=True), asi que nada evita que un segundo --commit
    duplique las 58 corridas con los mismos folios OP-2026-0001..0037. Este
    chequeo corre tanto en dry-run como en --commit para que el dry-run
    tambien reporte correctamente "esto ya se importo".

    No se puede usar el prefijo de folio OP-2026-% como senal: las ordenes
    en vivo usan el mismo esquema de folio, y cuando una orden en vivo se
    parte (ver _split_run_for_partial_material en
    backend/modules/production/service.py) su root_production_code queda en
    su propio OP-2026-NNNN. Eso daria un falso positivo y bloquearia el
    import aunque nunca se haya corrido. En cambio, production_run_event_lines
    solo se llena desde build_runs_for_order (mas abajo); ninguna orden en
    vivo escribe ahi, asi que su existencia es la senal correcta de "este
    import ya se corrio"."""
    from sqlalchemy import select

    existing = session.execute(
        select(ProductionRunEventLine.id).limit(1)
    ).first()
    if existing is not None:
        raise SystemExit(
            "Ya existen lineas de evento en production_run_event_lines, este "
            "import ya se corrio. Abortando para no duplicar."
        )


def build_runs_for_order(
    order: HistoricalOrder,
    folio: str,
    process: ProductionProcess,
    created_by_user_id: UUID,
) -> list[ProductionRun]:
    entrega_count = len(order.entrega_events)
    recibido_count = len(order.recibido_events)
    total = max(entrega_count, recibido_count, 1)

    runs: list[ProductionRun] = []
    for index in range(total):
        entrega = order.entrega_events[index] if index < entrega_count else None
        recibido = order.recibido_events[index] if index < recibido_count else None

        entrega_total = sum((line.gramos for line in entrega.lines), Decimal("0")) if entrega else Decimal("0")
        recibido_total = sum((line.gramos for line in recibido.lines), Decimal("0")) if recibido else Decimal("0")
        # raw_material_quantity_per_unit/unit_code no se usan para mostrar
        # (event_lines lo reemplaza), pero son NOT NULL: se guarda el total
        # entregado de esta corrida (o 1 si no hay entrega) para que
        # quantity=1 * eso siga siendo un total_required_material
        # internamente consistente. raw_material_item_id queda NULL — estas
        # ordenes no referencian ninguna materia prima real (decision #6).
        per_unit = entrega_total if entrega_total > 0 else Decimal("1")

        run = ProductionRun(
            id=uuid4(),
            process_id=process.id,
            process_name=order.order_name or f"Orden histórica {order.order_id}",
            quantity=Decimal("1"),
            status=ProductionRunStatus.RECEIVED if recibido else ProductionRunStatus.PENDING_RECEPTION,
            assembly_mode="ASIGNAR",
            raw_material_item_id=None,
            raw_material_quantity_per_unit=per_unit,
            raw_material_unit_code="g",
            total_required_material=per_unit,
            waste_limit_percent=process.waste_limit_percent,
            expected_finished_weight=per_unit,
            actual_finished_weight=recibido_total if recibido else None,
            production_code=folio if index == 0 else f"{folio}-{chr(ord('A') + index)}",
            root_production_code=folio,
            created_by_user_id=created_by_user_id,
            requested_at=dt.datetime.combine(
                (entrega.fecha if entrega and entrega.fecha else (recibido.fecha if recibido and recibido.fecha else dt.date(2026, 1, 1))),
                dt.time(9, 0),
            ),
        )
        if entrega:
            run.materials_approved_at = dt.datetime.combine(entrega.fecha, dt.time(9, 0)) if entrega.fecha else run.requested_at
            run.materials_approved_responsable_name = order.responsable
            for line_order, line in enumerate(entrega.lines):
                run.event_lines.append(
                    ProductionRunEventLine(
                        side="ENTREGA",
                        gramos=line.gramos,
                        unidad=raw_material.unit_code,
                        detalle=line.detalle,
                        line_order=line_order,
                    )
                )
        if recibido:
            run.received_at = dt.datetime.combine(recibido.fecha, dt.time(9, 0)) if recibido.fecha else run.requested_at
            run.received_responsable_name = order.responsable
            for line_order, line in enumerate(recibido.lines):
                run.event_lines.append(
                    ProductionRunEventLine(
                        side="RECEPCION",
                        gramos=line.gramos,
                        unidad=raw_material.unit_code,
                        detalle=line.detalle,
                        line_order=line_order,
                    )
                )
        run.stages = [
            ProductionRunStage(
                source_stage_id=stage.id,
                stage_name=stage.name,
                stage_type=stage.stage_type,
                stage_order=stage.stage_order,
                status=ProductionRunStageStatus.FINISHED,
                stage_code=f"{run.production_code}-{stage.stage_order}",
            )
            for stage in sorted(process.stages, key=lambda s: s.stage_order)
        ]
        runs.append(run)
    return runs


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--xlsx", type=Path, required=True)
    parser.add_argument("--created-by-username", type=str, required=True)
    parser.add_argument("--commit", action="store_true", help="Sin esta flag corre en modo dry-run.")
    args = parser.parse_args()

    orders = parse_orders(args.xlsx)
    print(f"Ordenes parseadas del Excel: {len(orders)}")

    session = SessionLocal()
    try:
        _abort_if_already_imported(session)
        user = _resolve_user(session, args.created_by_username)
        process = _get_or_create_process(session)
        folios = _next_folio_numbers(len(orders))

        all_runs: list[ProductionRun] = []
        for order, folio in zip(orders, folios):
            runs = build_runs_for_order(order, folio, process, user.id)
            all_runs.extend(runs)
            entrega_n = len(order.entrega_events)
            recibido_n = len(order.recibido_events)
            print(
                f"  {folio}: orden Excel #{order.order_id} '{order.order_name}' — "
                f"{len(runs)} corrida(s) ({entrega_n} entregas, {recibido_n} recepciones)"
            )

        print()
        print(f"Proceso: '{process.name}' ({process.id})")
        print(f"Usuario: '{user.username}' ({user.id})")
        print(f"Total corridas a insertar: {len(all_runs)}")
        print(f"Rango de folios raiz: {folios[0]} .. {folios[-1]}")

        if not args.commit:
            print()
            print("Modo dry-run: no se escribio nada. Corre de nuevo con --commit para confirmar.")
            session.rollback()
            return

        for run in all_runs:
            session.add(run)
        session.commit()
        print()
        print(f"Listo: {len(all_runs)} corridas insertadas.")
    finally:
        session.close()


if __name__ == "__main__":
    main()
